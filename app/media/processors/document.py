"""
Document processing module for PDF thumbnails, text extraction, and conversion.

Provides processing for documents:
- PDF: Direct thumbnail and text extraction
- Word (DOCX): Convert to PDF via LibreOffice, then process
- Excel (XLSX): Convert to PDF via LibreOffice, then process

Uses:
- pdf2image (poppler-utils) for PDF page rendering
- pdfplumber for text extraction
- LibreOffice headless for Office document conversion
- Pillow for image processing

Functions:
    extract_document_metadata: Extract page count, author, title
    convert_to_pdf: Convert Office documents to PDF
    generate_document_thumbnail: Create thumbnail of first page
    extract_document_text: Extract searchable text content
"""

from __future__ import annotations

import logging
import re
import subprocess
import tempfile
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any

from django.core.files.base import ContentFile
from PIL import Image

from media.processors.base import (
    DOCUMENT_CONVERSION_TIMEOUT,
    DOCUMENT_THUMBNAIL_DPI,
    DOCUMENT_THUMBNAIL_SIZE,
    MAX_TEXT_EXTRACTION_PAGES,
    PermanentProcessingError,
    TransientProcessingError,
    WEBP_QUALITY,
)

if TYPE_CHECKING:
    from media.models import MediaAsset, MediaFile

logger = logging.getLogger(__name__)


# =============================================================================
# Constants
# =============================================================================

# Supported document MIME types for conversion
OFFICE_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # .docx
    "application/msword",  # .doc
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls
    "application/vnd.openxmlformats-officedocument.presentationml.presentation",  # .pptx
    "application/vnd.ms-powerpoint",  # .ppt
    "application/vnd.oasis.opendocument.text",  # .odt
    "application/vnd.oasis.opendocument.spreadsheet",  # .ods
}

PDF_MIME_TYPE = "application/pdf"


# =============================================================================
# Exceptions
# =============================================================================


class DocumentProcessingError(PermanentProcessingError):
    """
    Raised when document processing fails permanently.

    This exception indicates a non-recoverable error such as:
    - Corrupted document file
    - Unsupported document format
    - Password-protected/encrypted document
    - Malformed PDF structure

    Tasks should NOT retry when this exception is raised.
    """

    pass


# =============================================================================
# Metadata Extraction
# =============================================================================


def extract_document_metadata(media_file: "MediaFile") -> dict[str, Any]:
    """
    Extract metadata from a document file.

    For PDFs extracts:
    - page_count
    - title, author, creator, producer
    - is_encrypted

    For Office documents extracts:
    - page_count (estimated)
    - title, author
    - sheet_count (for spreadsheets)

    Also determines:
    - is_searchable: Has extractable text
    - is_scanned: Appears to be scanned images (no text)

    Args:
        media_file: MediaFile instance with media_type='document'.

    Returns:
        Dictionary containing extracted metadata.

    Raises:
        DocumentProcessingError: If document cannot be read (permanent failure).
        TransientProcessingError: For timeout errors that should be retried.

    Example:
        >>> metadata = extract_document_metadata(media_file)
        >>> print(metadata)
        {
            'page_count': 15,
            'title': 'Quarterly Report',
            'author': 'John Doe',
            'is_searchable': True,
            'is_scanned': False,
            'is_encrypted': False,
            'format': 'pdf'
        }
    """
    logger.info(
        "Extracting metadata from document",
        extra={"media_file_id": str(media_file.pk)},
    )

    mime_type = media_file.mime_type

    if mime_type == PDF_MIME_TYPE:
        return _extract_pdf_metadata(media_file)
    elif mime_type in OFFICE_MIME_TYPES:
        return _extract_office_metadata(media_file)
    else:
        # Treat as generic document
        return {
            "format": _get_format_from_mime(mime_type),
            "is_searchable": False,
            "is_scanned": False,
            "is_encrypted": False,
        }


def _extract_pdf_metadata(media_file: "MediaFile") -> dict[str, Any]:
    """Extract metadata from a PDF file using pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        logger.error("pdfplumber not installed")
        raise TransientProcessingError(
            "pdfplumber not installed - required for PDF processing"
        )

    file_path = media_file.file.path
    metadata: dict[str, Any] = {
        "format": "pdf",
        "is_encrypted": False,
        "is_searchable": True,
        "is_scanned": False,
    }

    try:
        with pdfplumber.open(file_path) as pdf:
            # Page count
            metadata["page_count"] = len(pdf.pages)

            # PDF info (title, author, etc.)
            if pdf.metadata:
                info = pdf.metadata
                if info.get("Title"):
                    metadata["title"] = str(info["Title"])
                if info.get("Author"):
                    metadata["author"] = str(info["Author"])
                if info.get("Creator"):
                    metadata["creator"] = str(info["Creator"])
                if info.get("Producer"):
                    metadata["producer"] = str(info["Producer"])

            # Check if searchable by extracting text from first few pages
            text_found = False
            pages_to_check = min(3, len(pdf.pages))

            for i in range(pages_to_check):
                page_text = pdf.pages[i].extract_text()
                if page_text and len(page_text.strip()) > 50:
                    text_found = True
                    break

            metadata["is_searchable"] = text_found
            metadata["is_scanned"] = not text_found and metadata["page_count"] > 0

    except Exception as e:
        error_str = str(e).lower()

        # Check for encryption
        if "encrypted" in error_str or "password" in error_str:
            logger.warning(
                "PDF is encrypted/password-protected",
                extra={"media_file_id": str(media_file.pk)},
            )
            metadata["is_encrypted"] = True
            metadata["is_searchable"] = False
            return metadata

        # Check for corruption
        if "invalid" in error_str or "corrupt" in error_str or "malformed" in error_str:
            logger.warning(
                "PDF appears corrupted",
                extra={"media_file_id": str(media_file.pk), "error": str(e)},
            )
            raise DocumentProcessingError(f"PDF file is corrupted: {e}") from e

        logger.error(
            "Error extracting PDF metadata",
            extra={"media_file_id": str(media_file.pk), "error": str(e)},
        )
        raise DocumentProcessingError(f"Failed to read PDF metadata: {e}") from e

    logger.info(
        "Extracted PDF metadata successfully",
        extra={
            "media_file_id": str(media_file.pk),
            "page_count": metadata.get("page_count"),
            "is_searchable": metadata.get("is_searchable"),
        },
    )

    return metadata


def _extract_office_metadata(media_file: "MediaFile") -> dict[str, Any]:
    """Extract metadata from Office documents (docx, xlsx)."""
    mime_type = media_file.mime_type
    file_path = media_file.file.path

    metadata: dict[str, Any] = {
        "format": _get_format_from_mime(mime_type),
        "is_encrypted": False,
        "is_searchable": True,
        "is_scanned": False,
    }

    # Handle Word documents
    if (
        "wordprocessingml" in mime_type
        or "msword" in mime_type
        or "opendocument.text" in mime_type
    ):
        try:
            from docx import Document

            doc = Document(file_path)
            core_props = doc.core_properties

            if core_props.title:
                metadata["title"] = core_props.title
            if core_props.author:
                metadata["author"] = core_props.author

            # Estimate page count (rough estimate based on paragraphs)
            # Word doesn't store exact page count in the file
            para_count = len(doc.paragraphs)
            estimated_pages = max(1, para_count // 20)  # Rough estimate
            metadata["page_count"] = estimated_pages

        except Exception as e:
            error_str = str(e).lower()
            if "encrypted" in error_str or "password" in error_str:
                metadata["is_encrypted"] = True
                return metadata
            logger.warning(
                "Could not extract Word metadata",
                extra={"media_file_id": str(media_file.pk), "error": str(e)},
            )
            # Continue without detailed metadata

    # Handle Excel documents
    elif (
        "spreadsheetml" in mime_type
        or "ms-excel" in mime_type
        or "opendocument.spreadsheet" in mime_type
    ):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)

            if wb.properties:
                if wb.properties.title:
                    metadata["title"] = wb.properties.title
                if wb.properties.creator:
                    metadata["author"] = wb.properties.creator

            metadata["sheet_count"] = len(wb.sheetnames)
            wb.close()

        except Exception as e:
            error_str = str(e).lower()
            if "encrypted" in error_str or "password" in error_str:
                metadata["is_encrypted"] = True
                return metadata
            logger.warning(
                "Could not extract Excel metadata",
                extra={"media_file_id": str(media_file.pk), "error": str(e)},
            )

    logger.info(
        "Extracted Office document metadata",
        extra={
            "media_file_id": str(media_file.pk),
            "format": metadata.get("format"),
        },
    )

    return metadata


def _get_format_from_mime(mime_type: str) -> str:
    """Convert MIME type to simple format name."""
    mime_to_format = {
        "application/pdf": "pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
        "application/msword": "doc",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
        "application/vnd.ms-excel": "xls",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation": "pptx",
        "application/vnd.ms-powerpoint": "ppt",
        "application/vnd.oasis.opendocument.text": "odt",
        "application/vnd.oasis.opendocument.spreadsheet": "ods",
    }
    return mime_to_format.get(mime_type, "document")


# =============================================================================
# PDF Conversion
# =============================================================================


def convert_to_pdf(media_file: "MediaFile") -> Path:
    """
    Convert an Office document to PDF using LibreOffice headless.

    Args:
        media_file: MediaFile instance with Office document.

    Returns:
        Path to the converted PDF file (in temp directory).

    Raises:
        DocumentProcessingError: If conversion fails (permanent failure).
        TransientProcessingError: For timeout errors that should be retried.

    Note:
        The caller is responsible for cleaning up the returned temp file.
    """
    logger.info(
        "Converting document to PDF",
        extra={"media_file_id": str(media_file.pk), "mime_type": media_file.mime_type},
    )

    if media_file.mime_type == PDF_MIME_TYPE:
        # Already a PDF, return the original path
        return Path(media_file.file.path)

    if media_file.mime_type not in OFFICE_MIME_TYPES:
        raise DocumentProcessingError(
            f"Unsupported document type for PDF conversion: {media_file.mime_type}"
        )

    file_path = Path(media_file.file.path)

    # Create temp directory for output
    temp_dir = tempfile.mkdtemp(prefix="libreoffice_")
    temp_path = Path(temp_dir)

    try:
        # Run LibreOffice headless conversion
        cmd = [
            "libreoffice",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(temp_path),
            str(file_path),
        ]

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=DOCUMENT_CONVERSION_TIMEOUT,
        )

        if result.returncode != 0:
            stderr = result.stderr.strip() if result.stderr else "Unknown error"
            logger.warning(
                "LibreOffice conversion failed",
                extra={
                    "media_file_id": str(media_file.pk),
                    "returncode": result.returncode,
                    "stderr": stderr[:500],
                },
            )
            raise DocumentProcessingError(f"PDF conversion failed: {stderr}")

        # Find the output PDF
        output_name = file_path.stem + ".pdf"
        output_path = temp_path / output_name

        if not output_path.exists():
            # Try to find any PDF in the output directory
            pdfs = list(temp_path.glob("*.pdf"))
            if pdfs:
                output_path = pdfs[0]
            else:
                raise DocumentProcessingError("PDF conversion produced no output file")

        logger.info(
            "Document converted to PDF successfully",
            extra={
                "media_file_id": str(media_file.pk),
                "output_path": str(output_path),
            },
        )

        return output_path

    except subprocess.TimeoutExpired:
        logger.warning(
            "LibreOffice conversion timed out",
            extra={
                "media_file_id": str(media_file.pk),
                "timeout": DOCUMENT_CONVERSION_TIMEOUT,
            },
        )
        raise TransientProcessingError(
            f"PDF conversion timed out after {DOCUMENT_CONVERSION_TIMEOUT} seconds"
        )

    except FileNotFoundError:
        logger.error(
            "LibreOffice not found - ensure it is installed",
            extra={"media_file_id": str(media_file.pk)},
        )
        raise TransientProcessingError(
            "LibreOffice not found - required for document conversion"
        )


# =============================================================================
# Thumbnail Generation
# =============================================================================


def generate_document_thumbnail(media_file: "MediaFile") -> "MediaAsset":
    """
    Generate a thumbnail of the first page of a document.

    For PDFs, renders the first page directly.
    For Office documents, converts to PDF first, then renders.

    Args:
        media_file: MediaFile instance with media_type='document'.

    Returns:
        MediaAsset instance containing the generated thumbnail.

    Raises:
        DocumentProcessingError: If document cannot be processed (permanent failure).
        TransientProcessingError: For timeout errors that should be retried.
    """
    from media.models import MediaAsset

    try:
        from pdf2image import convert_from_path
    except ImportError:
        logger.error("pdf2image not installed")
        raise TransientProcessingError(
            "pdf2image not installed - required for document thumbnails"
        )

    logger.info(
        "Generating thumbnail for document",
        extra={"media_file_id": str(media_file.pk)},
    )

    pdf_path = None
    cleanup_pdf = False

    try:
        # Get PDF path (convert if necessary)
        if media_file.mime_type == PDF_MIME_TYPE:
            pdf_path = Path(media_file.file.path)
        else:
            pdf_path = convert_to_pdf(media_file)
            cleanup_pdf = True

        # Convert first page to image
        try:
            images = convert_from_path(
                pdf_path,
                first_page=1,
                last_page=1,
                dpi=DOCUMENT_THUMBNAIL_DPI,
            )
        except Exception as e:
            error_str = str(e).lower()
            if "encrypted" in error_str or "password" in error_str:
                raise DocumentProcessingError(
                    "Cannot generate thumbnail for encrypted PDF"
                ) from e
            if "invalid" in error_str or "corrupt" in error_str:
                raise DocumentProcessingError(
                    f"PDF is corrupted or invalid: {e}"
                ) from e
            raise DocumentProcessingError(f"Failed to render PDF page: {e}") from e

        if not images:
            raise DocumentProcessingError("PDF has no pages to render")

        # Get first page image
        page_img = images[0]

        # Create thumbnail (maintains aspect ratio)
        page_img.thumbnail(DOCUMENT_THUMBNAIL_SIZE, Image.Resampling.LANCZOS)

        # Convert to RGB if necessary
        if page_img.mode not in ("RGB", "RGBA"):
            page_img = page_img.convert("RGB")

        # Save as WebP
        buffer = BytesIO()
        page_img.save(buffer, format="WEBP", quality=WEBP_QUALITY)
        buffer.seek(0)

        width, height = page_img.size
        file_size = buffer.getbuffer().nbytes

        # Create or update the thumbnail asset
        asset, created = MediaAsset.objects.update_or_create(
            media_file=media_file,
            asset_type=MediaAsset.AssetType.THUMBNAIL,
            defaults={
                "width": width,
                "height": height,
                "file_size": file_size,
            },
        )

        filename = f"thumb_{media_file.pk}.webp"
        asset.file.save(filename, ContentFile(buffer.read()), save=True)

        logger.info(
            "Generated document thumbnail successfully",
            extra={
                "media_file_id": str(media_file.pk),
                "asset_id": str(asset.pk),
                "size": f"{width}x{height}",
                "file_size": file_size,
                "is_new": created,
            },
        )

        return asset

    finally:
        # Cleanup temporary PDF if we created one
        if cleanup_pdf and pdf_path and pdf_path.exists():
            try:
                pdf_path.unlink()
                # Also remove the temp directory
                if pdf_path.parent.exists():
                    pdf_path.parent.rmdir()
            except OSError:
                pass


# =============================================================================
# Text Extraction
# =============================================================================


def extract_document_text(media_file: "MediaFile") -> "MediaAsset":
    """
    Extract searchable text content from a document.

    For PDFs, extracts text directly using pdfplumber.
    For Office documents, converts to PDF first, then extracts.

    Text is stored as a plain text file asset for search indexing.
    Large documents (100+ pages) are limited to first 50 pages.

    If extraction fails or produces no text, the document is
    likely scanned/image-based (OCR not implemented).

    Args:
        media_file: MediaFile instance with media_type='document'.

    Returns:
        MediaAsset instance containing the extracted text.

    Raises:
        DocumentProcessingError: If document cannot be processed (permanent failure).
        TransientProcessingError: For timeout errors that should be retried.
    """
    from media.models import MediaAsset

    try:
        import pdfplumber
    except ImportError:
        logger.error("pdfplumber not installed")
        raise TransientProcessingError(
            "pdfplumber not installed - required for text extraction"
        )

    logger.info(
        "Extracting text from document",
        extra={"media_file_id": str(media_file.pk)},
    )

    pdf_path = None
    cleanup_pdf = False

    try:
        # Get PDF path (convert if necessary)
        if media_file.mime_type == PDF_MIME_TYPE:
            pdf_path = Path(media_file.file.path)
        else:
            pdf_path = convert_to_pdf(media_file)
            cleanup_pdf = True

        # Extract text from PDF
        extracted_text = []

        try:
            with pdfplumber.open(pdf_path) as pdf:
                # Limit pages for very large documents
                pages_to_extract = min(len(pdf.pages), MAX_TEXT_EXTRACTION_PAGES)

                for i in range(pages_to_extract):
                    page_text = pdf.pages[i].extract_text()
                    if page_text:
                        extracted_text.append(page_text)

        except Exception as e:
            error_str = str(e).lower()
            if "encrypted" in error_str or "password" in error_str:
                raise DocumentProcessingError(
                    "Cannot extract text from encrypted PDF"
                ) from e
            raise DocumentProcessingError(
                f"Failed to extract text from PDF: {e}"
            ) from e

        # Combine all extracted text
        full_text = "\n\n".join(extracted_text)

        # Clean up the text
        full_text = _clean_extracted_text(full_text)

        if not full_text or len(full_text.strip()) < 10:
            # No meaningful text extracted - likely a scanned document
            logger.info(
                "No text extracted - document may be scanned/image-based",
                extra={"media_file_id": str(media_file.pk)},
            )
            # Update metadata to indicate scanned document
            if media_file.metadata:
                media_file.metadata["is_scanned"] = True
                media_file.metadata["is_searchable"] = False
                media_file.save(update_fields=["metadata"])

            raise DocumentProcessingError(
                "No text could be extracted - document may be scanned/image-based"
            )

        # Save as text file
        text_bytes = full_text.encode("utf-8")
        buffer = BytesIO(text_bytes)
        file_size = len(text_bytes)

        # Create or update the extracted text asset
        asset, created = MediaAsset.objects.update_or_create(
            media_file=media_file,
            asset_type=MediaAsset.AssetType.EXTRACTED_TEXT,
            defaults={
                "file_size": file_size,
            },
        )

        filename = f"text_{media_file.pk}.txt"
        asset.file.save(filename, ContentFile(buffer.read()), save=True)

        logger.info(
            "Extracted document text successfully",
            extra={
                "media_file_id": str(media_file.pk),
                "asset_id": str(asset.pk),
                "text_length": len(full_text),
                "file_size": file_size,
                "is_new": created,
            },
        )

        return asset

    finally:
        # Cleanup temporary PDF if we created one
        if cleanup_pdf and pdf_path and pdf_path.exists():
            try:
                pdf_path.unlink()
                if pdf_path.parent.exists():
                    pdf_path.parent.rmdir()
            except OSError:
                pass


def _clean_extracted_text(text: str) -> str:
    """
    Clean up extracted text for better searchability.

    - Removes excessive whitespace
    - Normalizes line endings
    - Removes control characters
    """
    if not text:
        return ""

    # Remove control characters except newlines and tabs
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)

    # Normalize line endings
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    # Remove excessive blank lines (more than 2)
    text = re.sub(r"\n{3,}", "\n\n", text)

    # Remove excessive spaces
    text = re.sub(r"[ \t]+", " ", text)

    # Strip leading/trailing whitespace from each line
    lines = [line.strip() for line in text.split("\n")]
    text = "\n".join(lines)

    return text.strip()
